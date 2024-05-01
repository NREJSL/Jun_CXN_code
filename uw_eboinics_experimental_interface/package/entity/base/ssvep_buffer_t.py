
import numpy as np
import joblib
import os
import mne
from PyQt5.QtWidgets import QMainWindow
from package.entity.base.buffer import Buffer
from package.views.layouts import online_test_layout
from package.entity.base.utils import SSVEPCCAAnalysis
import pdb
import pylsl
import serial
import serial.tools.list_ports
import time
from pandas import DataFrame
import pandas as pd
import openpyxl
from collections import Counter



com_port = 'COM3'  # Change this to the COM port you want to use
baud_rate = 115200 
temp=0
temp2=3

classification=[3,3,3]

#ser.timeout=1
data_to_send_dict = {0: 'L', 1: 'F', 2: 'R', 3: 'S'}
ser = serial.Serial(com_port, baud_rate)
coefficient_ls=[0.5,0.5,0.5,0,5]
class SSVEPBuffer(Buffer):
    '''
    SSVEPBuffer is a child class of Buffer, it handles online processing and online testing.

    Parameter
    ----------
    window_stride: (second) Time interval between two windows
    window_size: (second) length of time window
    buffer_size: (second) length of buffer in StreamReceiver
    l_filter: None or object of Filter class
    filter_type: 'bpf' ---- band pass filter
                 'lpf' ---- low pass filter
                 'hpf' ---- high pass filter
                 'bsf; ---- band stop filter (notch filter)
    downsample: None or int, downsample rate
    model_path: None or file path containing .pkl file
    model_type: 'sklearn' or 'keras'. Further updating could be added for each configuration.
    model_name: name of classifier. E.g. cca, etc.

    '''
    def __init__(self, main_view, window_stride=0.75, window_size=1.5, buffer_size=10,
                 l_filter=None, filter_type=None, ica_path=None, downsample=None, model_path=None,
                 model_type=None, model_name='cca', stimulus_type='ssvep', channels_list=['C1', 'C2', 'C3','C4', 'C5', 'C6'],
                 target_frequencies=None):
        super().__init__(window_stride, window_size, buffer_size, l_filter, filter_type)    
        self.main_view = main_view
        self.downsample = downsample
        self.ica_model = None
        self.stimulus_type = stimulus_type
        self.model_type = model_type
        self.model_name = model_name
        self.target_frequencies = target_frequencies
        self.target_frequencies = [9,7,11,13]
        self.cca_pipeline = 0
        self.predicted_class = None
        self.required_channels = channels_list
        self.channel_indexes = []
        
        corr_coeff=[0,0,0,0]
        self.coefficient_ls=[0.5,0.5,0.5,0,5]
        self.control= 0
        self.curClass=3

        wbook=openpyxl.Workbook(r"C:\Users\COSMEDUSER\Desktop\Test.xlsx")
        
        #workbook=openpyxl.Workbook()
        #sheet_active=workbook.active
        #sheet_active.title="Sheet1"
        #excel_path=r'C:\Users\COSMEDUSER\Desktop\Test1.xlsx'
        #workbook.save(excel_path)
        
        print(self.eeg_ch_names)
        for ch_name in self.required_channels:
            if ch_name in self.eeg_ch_names:
                self.channel_indexes.append(self.eeg_ch_names.index(ch_name))
                
        print('self.channel_indexes: ', self.channel_indexes)
    
        if downsample:
            self.sf = downsample
            self.n_sample = int(self.window_size * self.sf)
            self.window = np.zeros((self.n_ch, self.n_sample))
            self.eeg_window = np.zeros((self.n_eeg_ch, self.n_sample))

        if ica_path is not None:
            if os.path.exists(ica_path):
                self.ica_model = mne.preprocessing.read_ica(ica_path)
                
        if self.model_name=='cca':
            self.cca_pipeline = 1
        else:
            raise NotImplementedError('Reqeusted method not implemented')

        self.mne_info = mne.create_info(ch_names=self.eeg_ch_names, sfreq=self.sf, ch_types='eeg')
        self.mne_info.set_montage('standard_1020')
    
    def loop(self):
        '''
        Get called every window_stride seconds.
        It reads chunks of data using stream_receiver and process them.
        Processing steps include filter, downsample, ica, prediction.

        '''
        data, self.ts_list = self.sr.acquire('buffer using', blocking=True)
        if len(self.ts_list) > 0:
            data = data[:, self.sr.get_channels()].T / self.sr.multiplier
            # pdb.set_trace()
            self.window = np.roll(self.window, -len(self.ts_list), 1)
            if self.filter:
                filtered_data, _ = self.filter.apply_filter(data)
                self.window[:, -len(self.ts_list):] = filtered_data
            else:
                self.window[:, -len(self.ts_list):] = data
            
            self.eeg_window = self.window[self.eeg_ch_idx, :]

            eeg_window_3d = np.expand_dims(self.eeg_window, axis=0)
            window_epoch = mne.EpochsArray(eeg_window_3d, info=self.mne_info)
            # pdb.set_trace()
            if self.downsample:
                window_epoch = window_epoch.resample(self.downsample)
            if self.ica_model:
                window_epoch = self.ica_model.apply(window_epoch)
                # pdb.set_trace()
            
            # predict_window = window_epoch.get_data().copy()
            self.eeg_window = window_epoch.get_data()[0, :, :]
            if len(self.channel_indexes)==len(self.required_channels):
                self.detection_window = self.eeg_window[self.channel_indexes, :]
            else:
                raise NotImplementedError('Requested channels not in the channel list')
            
            if self.cca_pipeline:
                cca_analysis_object = SSVEPCCAAnalysis(fs=self.sf, data_len=self.window_size, 
                                                       target_freqs=self.target_frequencies, 
                                                       num_harmonics=2)
                corr_coeff = cca_analysis_object.apply_cca(self.detection_window.T)
                #average=sum(corr_coeff)/4
                lslclock=pylsl.local_clock()
               
                

            print(" ")
            print("Result")
            
            #print("previous LS",round(coefficient_ls[0],4),round(coefficient_ls[1],4),round(coefficient_ls[2],4),round(coefficient_ls[3],4))


            #print(round(corr_coeff[0],4),round(corr_coeff[1],4),round(corr_coeff[2],4),round(corr_coeff[3],4))
            
            '''
            step=0.5
            # classification rate limiting
            if(corr_coeff[0]-coefficient_ls[0]>step):
                coefficient_ls[0]=coefficient_ls[0] + 1.2*step
            elif(corr_coeff[0]-coefficient_ls[0]< -1*step):
                coefficient_ls[0]=coefficient_ls[0] - 1*step
            else:
                coefficient_ls[0]=corr_coeff[0]

            if(corr_coeff[1]-coefficient_ls[1]>step):
                coefficient_ls[1]=coefficient_ls[1] + 1.2*step
            elif(corr_coeff[1]-coefficient_ls[1]< -1*step):
                coefficient_ls[1]=coefficient_ls[1] - 1*step
            else:
                coefficient_ls[1]=corr_coeff[1]-0.05

            if(corr_coeff[2]-coefficient_ls[2]>step):
                coefficient_ls[2]=coefficient_ls[2] + 1.2* step
            elif(corr_coeff[2]-coefficient_ls[2]< -1*step):
                coefficient_ls[2]=coefficient_ls[2] - 1*step
            else:
                coefficient_ls[2]=corr_coeff[2]

            print("Current LS",round(coefficient_ls[0],4),round(coefficient_ls[1],4),round(coefficient_ls[2],4),round(coefficient_ls[3],4))
            '''


            thres=0.65
            #coefficient_ls[1]=coefficient_ls[1]
            '''
            if(max(corr_coeff[0:3])>thres):
           
                self.predicted_class = np.argmax(corr_coeff[0:3], axis=-1)
               
            else:
                self.predicted_class = 3
            '''

            ###################################
            #Toggle control
            '''
            if(max(corr_coeff[0:2])>thres and self.control == 1):
                self.curClass = np.argmax(corr_coeff[0:2], axis=-1)
                self.control=0
            
            elif(corr_coeff[3]>thres and self.control == 0):
                self.control = 1
            else:
                self.predicted_class = self.curClass
              '''
            if(corr_coeff[3]>0.5):
                self.control=1

            if(self.control == 1):
                ser.write(bytes(data_to_send_dict[3], encoding='utf-8'))
                 
                if(max(corr_coeff[0:2])>thres):
                    self.curClass = np.argmax(corr_coeff[0:2], axis=-1)
                    
                    ser.write(bytes(data_to_send_dict[self.curClass], encoding='utf-8'))
                    time.sleep(0.001)
                    self.control=0
            elif(self.control == 0):
                ser.write(bytes(data_to_send_dict[self.curClass], encoding='utf-8'))
            print("-------------------------------Sent byte is : ", self.curClass)


                
            print("-----------------Toggle Status: ", self.control)

            '''
                corr_mod=corr_coeff
            
            corr_mod[0]=corr_mod[0]-min(corr_coeff)
            corr_mod[1]=corr_mod[1]-min(corr_coeff)
            corr_mod[2]=corr_mod[2]-min(corr_coeff)
            print(corr_mod)
            if(max(corr_mod)>0.1):
                self.predicted_class=np.argmax(corr_mod[0:2],axis=-1)
            else:
                self.predicted_class=3

            
            '''  
            '''
            if(sum(corr_coeff)/4 <0.35):
                self.predicted_class=self.temp2 
                self.temp=self.temp2   
                #print(round(corr_coeff[0],4),round(corr_coeff[1],4),round(corr_coeff[2],4),"Low Average Override")        
            elif((corr_coeff[1])>0.6 and corr_coeff[1]==max(corr_coeff)):
                self.predicted_class = 1
                self.temp=1
                #print(round(corr_coeff[0],4),round(corr_coeff[1],4),round(corr_coeff[2],4),"Loop F")
            elif((corr_coeff[2])>0.6 and corr_coeff[2]==max(corr_coeff)):
                self.predicted_class = 2
                self.temp=2
                #print(round(corr_coeff[0],4),round(corr_coeff[1],4),round(corr_coeff[2],4),"Loop R"  )
            elif((corr_coeff[0])>0.6 and corr_coeff[0]==max(corr_coeff)):
                self.predicted_class = 0
                self.temp=0
                #print(round(corr_coeff[0],4),round(corr_coeff[1],4),round(corr_coeff[2],4),"Loop L")
                            
            else:
                self.predicted_class=3
                #print(round(corr_coeff[0],4),round(corr_coeff[1],4),round(corr_coeff[2],4),"No classification" )
                self.temp=3
            '''  
            #timestamp = time.strftime('%Y%m%d-%H%M%S', time.localtime())
            




            ########################################
            #Double verification
            '''
            print("Current classification------------------------", self.predicted_class)
            classification[2]=self.predicted_class
            classification[1]=self.repeat_most(classification)
            '''
            '''
            i=0
            while(i<1):
                #ser.write(bytes(data_to_send_dict[classification[1]], encoding='utf-8'))
                ser.write(bytes(data_to_send_dict[self.curClass], encoding='utf-8'))
                #time.sleep(0.01)
                #time.sleep(0.01)
                i=i+1
            '''
            #print("-------------------------------Sent byte is : ", classification[1])
            #print("-------------------------------Sent byte is : ", self.predicted_class)
            
            
            corr_time=corr_coeff.copy()
            corr_time.insert(0, classification[1])
            corr_time.insert(0, thres)
            corr_time.insert(0, lslclock)
            

            try:
                wb= openpyxl.load_workbook(r"C:\Users\COSMEDUSER\Desktop\Test.xlsx")
                ws=wb.active
                ws.append(corr_time)
                wb.save(r"C:\Users\COSMEDUSER\Desktop\Test.xlsx")
            except:
                print("Excel Error")



            print(round(corr_coeff[0],4),round(corr_coeff[1],4),round(corr_coeff[2],4),round(corr_coeff[3],4))
            #print(round(coefficient_ls[0],4),round(coefficient_ls[1],4),round(coefficient_ls[2],4),round(coefficient_ls[3],4))
            #print("Decision ", classification, " ", classification[1])

            
           # classification[0]=classification[1]
           # classification[1]=classification[2]
            
            
            self.window[self.eeg_ch_idx, :] = self.eeg_window

    def update_window_stride(self, new_stride):
        '''
        Update window stride on the fly
        :param new_stride: int
        '''
        self.stop_timer()
        self.start_timer()
        self.window_stride = new_stride

    def repeat_most(self, decision):
        counts=Counter(decision)
        most_common= counts.most_common(3)
        if len(most_common)>1 and most_common[0][1]==most_common[1][1]:
            return 3
        else:
            return most_common[0][0] if most_common else None